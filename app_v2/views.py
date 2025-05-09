import json
import random
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from datetime import date, datetime
from django.contrib import messages
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.db.models import Count
from django.views.decorators.csrf import csrf_exempt


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import User, Leader
import json

@csrf_exempt
def assign_leader(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        user_id = data.get('user_id')
        leader_id = data.get('leader_id')

        try:
            user = User.objects.get(id=user_id)
            leader = Leader.objects.get(id=leader_id)
            
            # Update user's leader
            user.leader = leader
            user.save()

            return JsonResponse({'status': 'success'})
        except User.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=400)
        except Leader.DoesNotExist:
            return JsonResponse({'error': 'Leader not found'}, status=400)

def index(request):
    if request.method == 'POST':
        gamename = request.POST.get('game_name')
        print(gamename)
        request.session['name'] = gamename
        return redirect('add_user')
    quizzes = Quiz.objects.all()
    today = date.today().isoformat()
    return render(request, "index.html", {'quizzes': quizzes, 'today_date': today})

def delete_user(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        user.delete()
       
        return redirect('add_user') 
    return HttpResponse("Invalid request method.")

@csrf_exempt
def update_leader_null(request):
    if request.method == 'POST':
        # Бүх хэрэглэгчийн удирдагчийг null болгох
        users = User.objects.all()
        for user in users:
            user.leader = None
            user.save()

        return redirect('team_count')

def team_count(request):
    leaders = Leader.objects.all().order_by('order')  # ⬅ дарааллаар нь сортлох
    leader_map = {}
    for leader in leaders:
        leader_map[leader] = User.objects.filter(leader=leader)

    unassigned_users = User.objects.filter(leader=None)
    return render(request, 'team_count.html', {
        'leader_map': leader_map,
        'unassigned_users': unassigned_users
    })

def shuffle_leader_order(request):
    if request.method == 'POST':
        leaders = list(Leader.objects.all())
        random.shuffle(leaders)

        for i, leader in enumerate(leaders):
            leader.order = i + 1
            leader.save()

        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'invalid'}, status=400)

def adduser(request):
    users = User.objects.all()

    if request.method == 'POST':
        username = request.POST.get('username')

        if User.objects.filter(username=username).exists():
            messages.error(request, "This username is already taken. Please choose a different one.")
            return redirect('add_user')

        if username:
            try:
                # Create a new user and assign them to the selected team
                user = User(username=username)
                user.save()
                messages.success(request, f"User '{username}' added successfully.")
            except Exception as e:
                # Catch any errors while saving the user
                messages.error(request, f"Error adding user: {str(e)}")
        else:
            messages.error(request, "Username is required.")
        
        return redirect('add_user')  # Redirect back to the form

    # Get the list of users to display on the page
    return render(request, "adduser.html", {'users': users})

def quiz(request):
    quizzes = Quiz.objects.all()
    
    # Get selected quiz ID from GET request
    quiz_id = request.GET.get('quiz_id')
    
    # If a quiz is selected, get its questions
    selected_quiz = None
    if quiz_id:
        selected_quiz = Quiz.objects.get(id=quiz_id)
        subquizzes = SubQuiz.objects.filter(quiz=selected_quiz)
    else:
        subquizzes = []
    
    return render(
        request, 
        'quiz.html', 
        {
            'quizzes_data': quizzes, 
            'selected_quiz': selected_quiz, 
            'subquizzes': subquizzes
        }
    )

def subquiz_detail(request, subquiz_id):
    subquiz = get_object_or_404(SubQuiz, id=subquiz_id)
    return render(
        request,
        'subquiz_detail.html',
        {
            'subquiz': subquiz
        }
    )

def show_time(request, subquiz_id):
    quiz = SubQuiz.objects.get(id=subquiz_id)
    return render(request, 'show_time.html', {'quiz': quiz})

from django.shortcuts import render
from .models import Leader, User, SubQuiz, Point
from django.db.models import Sum

def point(request):
    if request.method == 'POST':
        subquiz_id = request.GET.get('subquiz')

        if subquiz_id:
            try:
                subquiz = SubQuiz.objects.get(id=subquiz_id)

                # Зөв
                if request.POST.get('leader_id_yes'):
                    leader = Leader.objects.get(id=request.POST.get('leader_id_yes'))
                    Point.objects.create(
                        leader=leader,
                        subquiz=subquiz,
                        point=subquiz.point
                    )

                # Буруу
                if request.POST.get('leader_id_no'):
                    leader = Leader.objects.get(id=request.POST.get('leader_id_no'))
                    Point.objects.create(
                        leader=leader,
                        subquiz=subquiz,
                        point=0
                    )

            except SubQuiz.DoesNotExist:
                pass
            except Leader.DoesNotExist:
                pass

    leaders = Leader.objects.all().order_by('order')
    leader_map = {leader: User.objects.filter(leader=leader) for leader in leaders}
    
    # 🟢 Leader бүрийн оноог цуглуулах
    leader_scores = {
        leader.id: Point.objects.filter(leader=leader).aggregate(total=Sum('point'))['total'] or 0
        for leader in leaders
    }

    return render(request, 'point.html', {
        'leader_map': leader_map,
        'leader_scores': leader_scores,
    })

def endgame(request):
    # Бүх удирдагч нарыг авна
    leaders = Leader.objects.all()

    # 🟢 Удирдагч бүрийн оноог цуглуулах
    leader_scores = {}
    for leader in leaders:
        total = Point.objects.filter(leader=leader).aggregate(total=Sum('point'))['total'] or 0
        leader_scores[leader.id] = total

    # 🔵 Оноогоор буурч эрэмбэлсэн удирдагчдын жагсаалт
    sorted_leaders = sorted(leaders, key=lambda l: leader_scores[l.id], reverse=True)

    # 🥇 Хамгийн өндөр оноотой баг
    top_leader = sorted_leaders[0] if sorted_leaders else None
    top_score = leader_scores[top_leader.id] if top_leader else 0

    return render(request, 'endgame.html', {
        'leaders': sorted_leaders,
        'leader_scores': leader_scores,
        'top_leader': top_leader,
        'top_score': top_score,
    })

def reset(request):
    # Бүх оноог устгах
    Point.objects.all().delete()

    # Бүх удирдагчийн оноог 0 болгох
    leaders = Leader.objects.all()
    for leader in leaders:
        leader.point = 0
        leader.save()

    return redirect('index')

def save_data_to_excel(request):
    # Retrieve all teams, their users, and scores
    teams = Team.objects.all()
    game_name = request.session.get('name', 'Game')  # Default to 'Game' if not in session
    
    # Generate folder name based on today's date
    today_date = datetime.now().strftime('%Y-%m-%d')  # Format: YYYY-MM-DD
    directory_path = os.path.join('game_data', today_date)  # Create a folder like 'game_data/2024-12-03'

    # Make sure the directory exists
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

    # Generate the file name with timestamp
    timestamp = datetime.now().strftime('%H-%M-%S')  # Format: HH-MM-SS for the file name
    file_path = os.path.join(directory_path, f'{game_name}_{today_date}_{timestamp}.xlsx')

    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Game Data"

    # Create Header Row
    sheet.merge_cells('A1:C1')
    header_cell = sheet.cell(row=1, column=1)
    header_cell.value = f"{game_name} - {today_date}"
    header_cell.font = Font(size=14, bold=True)
    header_cell.alignment = Alignment(horizontal='center')

    # Add column headers
    sheet.append(["Багийн нэр", "Гишүүдийн нэрс", "Оноо"])

    # Fill in body with team data
    highest_score = 0
    best_team_name = None
    tie_breaker_data = {}  # To store data for tiebreaker logic

    for team in teams:
        # Retrieve all users for the current team
        users = User.objects.filter(team=team)
        user_names = ", ".join([user.username for user in users]) if users.exists() else "Гишүүнгүй."

        # Retrieve all scores for the team and ensure ordering
        scores = Score.objects.filter(team=team).order_by('-id')  # Ensure the queryset is ordered
        most_recent_score = scores.first()  # Get the most recent score
        score_value = most_recent_score.score if most_recent_score else 0

        # Track the highest score and corresponding team
        if score_value >= highest_score:
            # Store tiebreaker data
            highest_score_count = scores.values('score').annotate(count=Count('score')).filter(score=highest_score).first()
            tie_breaker_data[team.temaname] = highest_score_count['count'] if highest_score_count else 0
            
            if score_value > highest_score:
                highest_score = score_value
                best_team_name = team.temaname
            elif score_value == highest_score:
                # Resolve ties based on the highest score occurrence
                if tie_breaker_data[team.temaname] > tie_breaker_data.get(best_team_name, 0):
                    best_team_name = team.temaname

        # Write team data
        sheet.append([team.temaname, user_names, score_value])


    # Add footer row for highest scoring team
    footer_start = sheet.max_row + 1
    sheet.merge_cells(f'A{footer_start}:C{footer_start}')
    footer_cell = sheet.cell(row=footer_start, column=1)
    footer_cell.value = f"Хамгийн өндөр оноотой баг: {best_team_name}"
    footer_cell.font = Font(size=12, bold=True, color="FF0000")
    footer_cell.alignment = Alignment(horizontal='center')

    # Save the Excel file
    workbook.save(file_path)

    # After saving, provide feedback to the user
    return redirect('index')
